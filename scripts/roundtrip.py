#!/usr/bin/env -S python3 -u
"""Two-part roundtrip test: read → JSON → write → verify.

Usage:
    python scripts/roundtrip.py Data/Example.xlsx Data/Empty.xlsx
    python scripts/roundtrip.py Data/Empty.xlsx Data/Empty.xlsx

First argument is the source workbook to read.
Second argument is the blank template to write into.
Results are saved to records/roundtrip_<timestamp>/.

PART 1 — No Excel required (pure openpyxl):
  Phase 1 (read inputs): openpyxl reads input cells only (skip_formulas=True).
  Phase 2 (read all):    openpyxl reads ALL cells (skip_formulas=False) for
                         formula filter statistics.
  Phase 3 (write):       openpyxl writes inputs into the template.
  Phase 4 (verify):      openpyxl reads the written file at the writer's
                         cell addresses and verifies every written cell.

PART 2 — Excel required (xlwings, optional):
  Phase 5 (Excel read):  Read the original source via xlwings+Excel to
                         compare live-recalculated values against openpyxl's
                         cached values from Part 1.

Part 2 is skipped if xlwings or Excel is not available. Part 1 alone
confirms input data fidelity without any Excel dependency.

See also: scripts/verify_excel.py — post-Excel full-fidelity comparison
(run after manually opening both files in Excel and saving).
"""

from __future__ import annotations

import json
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from phpp_tool.locators import col_to_idx
from phpp_tool.reader import read_phpp
from phpp_tool.writer import write_phpp

FIELD_MAP = str(ROOT / "phpp-field-mapping.md")


def _count_values(data: dict, depth: int = 0) -> tuple[int, int]:
    """Count (total_values, non_none_values) recursively."""
    total = 0
    non_none = 0
    for v in data.values():
        if isinstance(v, dict):
            t, n = _count_values(v, depth + 1)
            total += t
            non_none += n
        elif isinstance(v, list):
            for item in v:
                if isinstance(item, dict):
                    t, n = _count_values(item, depth + 1)
                    total += t
                    non_none += n
                else:
                    total += 1
                    if item is not None:
                        non_none += 1
        else:
            total += 1
            if v is not None:
                non_none += 1
    return total, non_none


def _values_match(expected: Any, actual: Any) -> bool:
    """Compare values with tolerance for float rounding."""
    if expected == actual:
        return True
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        if abs(expected) < 1e-10 and abs(actual) < 1e-10:
            return True
        if abs(expected) > 0:
            return abs(expected - actual) / abs(expected) < 1e-6
    if isinstance(expected, float) and isinstance(actual, int):
        return _values_match(expected, float(actual))
    if isinstance(expected, int) and isinstance(actual, float):
        return _values_match(float(expected), actual)
    return False


def _verify_writes(
    written_path: Path,
    writes: list[tuple[str, str, int, Any]],
) -> tuple[int, int, list[str]]:
    """Verify every cell the writer targeted (openpyxl, no Excel)."""
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)
    from openpyxl import load_workbook

    wb = load_workbook(str(written_path), data_only=False)
    checked = 0
    mismatched = 0
    details: list[str] = []

    for sheet_name, col, row, expected in writes:
        if sheet_name not in wb.sheetnames:
            mismatched += 1
            details.append(f"  MISSING SHEET: {sheet_name}")
            continue

        ws = wb[sheet_name]
        actual = ws.cell(row=row, column=col_to_idx(col)).value
        checked += 1

        if not _values_match(expected, actual):
            mismatched += 1
            details.append(
                f"  {sheet_name}!{col}{row}:"
                f" expected {expected!r} ({type(expected).__name__}),"
                f" got {actual!r} ({type(actual).__name__})")

    wb.close()
    return checked, mismatched, details


# ======================================================================
# Part 2: Excel-based full comparison (optional)
# ======================================================================

def _xlwings_full_read(path: Path, field_map: str) -> dict[str, Any] | None:
    """Read ALL cells via xlwings+Excel (skip_formulas=False).

    Returns None if xlwings or Excel is not available.
    """
    try:
        import xlwings as xw
    except ImportError:
        return None

    try:
        from phpp_tool.locators import prefer_si_sheet
        from phpp_tool.map_parser import parse_field_map

        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
    except Exception:
        return None

    try:
        wb = app.books.open(str(path.resolve()))
        fm = parse_field_map(field_map)
        sheet_names = [s.name for s in wb.sheets]
        result: dict[str, Any] = {}

        for ws_key, ws_spec in fm.items():
            sheet_name = prefer_si_sheet(ws_spec["sheet_name"], sheet_names)
            if sheet_name not in sheet_names:
                continue
            ws = wb.sheets[sheet_name]
            ws_result: dict[str, Any] = {}

            for field_name, spec in ws_spec.get("fields", {}).items():
                if not spec.get("locator_string"):
                    continue
                from phpp_tool.locators import norm, find_row_in_col as _fric
                loc_col = spec["locator_col"]
                needle = norm(spec["locator_string"])
                col_idx = col_to_idx(loc_col)
                last_row = ws.used_range.last_cell.row
                found_row = None
                vals = ws.range((1, col_idx), (last_row, col_idx)).value
                if vals is None:
                    continue
                if not isinstance(vals, list):
                    vals = [vals]
                for i, cv in enumerate(vals):
                    if cv and needle in norm(cv):
                        found_row = 1 + i + spec.get("row_offset", 0)
                        break
                if found_row is None:
                    continue
                input_idx = col_to_idx(spec["input_col"])
                val = ws.range((found_row, input_idx)).value
                if val is not None:
                    ws_result[field_name] = val

            if ws_result:
                result[ws_key] = ws_result

        wb.close()
    finally:
        app.quit()

    return result


def _flatten_for_compare(data: dict[str, Any]) -> dict[str, Any]:
    """Flatten to top-level scalar fields per worksheet."""
    flat: dict[str, Any] = {}
    for ws_key, ws_data in data.items():
        if not isinstance(ws_data, dict):
            continue
        ws_flat = {k: v for k, v in ws_data.items()
                   if v is not None and not isinstance(v, (dict, list))}
        if ws_flat:
            flat[ws_key] = ws_flat
    return flat


def _compare_full(
    orig_data: dict[str, Any],
    written_data: dict[str, Any],
) -> tuple[int, int, list[str]]:
    """Compare all fields between two full-reads."""
    checked = 0
    mismatched = 0
    details: list[str] = []

    all_keys = set(orig_data.keys()) | set(written_data.keys())
    for ws_key in sorted(all_keys):
        orig_ws = orig_data.get(ws_key, {})
        written_ws = written_data.get(ws_key, {})
        all_fields = set(orig_ws.keys()) | set(written_ws.keys())
        for field in sorted(all_fields):
            orig_val = orig_ws.get(field)
            written_val = written_ws.get(field)
            if orig_val is None and written_val is None:
                continue
            checked += 1
            if not _values_match(orig_val, written_val):
                mismatched += 1
                details.append(
                    f"  {ws_key}.{field}:"
                    f" original={orig_val!r}, written={written_val!r}")

    return checked, mismatched, details


# ======================================================================
# Main roundtrip
# ======================================================================

def roundtrip(source: Path, template: Path, out_dir: Path) -> dict:
    source_name = source.stem
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'=' * 60}")
    print(f"  ROUNDTRIP: {source.name} → {template.name}")
    print(f"{'=' * 60}")

    # ==================================================================
    # PART 1 — Pure openpyxl (no Excel)
    # ==================================================================
    print(f"\n  PART 1: Pure openpyxl verification")
    print(f"  {'─' * 40}")

    # Step 1: read source — input cells only
    print(f"\n[1/4] Reading {source.name} — inputs only ...")
    t0 = time.time()
    data_inputs = read_phpp(str(source), FIELD_MAP, skip_formulas=True)
    t_read_inputs = time.time() - t0
    total_i, non_none_i = _count_values(data_inputs)
    print(f"      {len(data_inputs)} worksheets,"
          f" {non_none_i} non-None / {total_i} total values,"
          f" {t_read_inputs:.1f}s")

    json_inputs = out_dir / f"{source_name}_inputs.json"
    json_inputs.write_text(
        json.dumps(data_inputs, indent=2, default=str), encoding="utf-8"
    )

    # Step 2: read source — all cells (for formula filter stats)
    print(f"\n[2/4] Reading {source.name} — all cells ...")
    t0 = time.time()
    data_all = read_phpp(str(source), FIELD_MAP, skip_formulas=False)
    t_read_all = time.time() - t0
    total_a, non_none_a = _count_values(data_all)
    filtered = non_none_a - non_none_i
    pct = 100 * filtered / max(non_none_a, 1)
    print(f"      {len(data_all)} worksheets,"
          f" {non_none_a} non-None / {total_a} total values,"
          f" {t_read_all:.1f}s")
    print(f"      Formula filter removed {filtered} values ({pct:.1f}%)")

    json_all = out_dir / f"{source_name}_all.json"
    json_all.write_text(
        json.dumps(data_all, indent=2, default=str), encoding="utf-8"
    )

    # Step 3: write into template
    written_path = out_dir / f"{source_name}_written.xlsx"
    print(f"\n[3/4] Writing into {template.name} → {written_path.name} ...")
    t0 = time.time()
    writes = write_phpp(data_inputs, str(template), str(written_path),
                        FIELD_MAP)
    t_write = time.time() - t0
    print(f"      {len(writes)} cell writes, {t_write:.1f}s")

    # Step 4: verify every written cell
    print(f"\n[4/4] Verifying {len(writes)} written cells ...")
    t0 = time.time()
    checked, mismatched, mismatch_details = _verify_writes(
        written_path, writes)
    t_verify = time.time() - t0

    if mismatched == 0:
        print(f"      *** All {checked} cells verified — PERFECT MATCH ***")
    else:
        print(f"      {checked} checked, {mismatched} MISMATCHES:")
        for d in mismatch_details[:20]:
            print(d)
        if len(mismatch_details) > 20:
            print(f"      ... and {len(mismatch_details) - 20} more")
    print(f"      {t_verify:.1f}s")

    part1_t = t_read_inputs + t_read_all + t_write + t_verify
    print(f"\n  Part 1 timing: {part1_t:.1f}s total")

    # ==================================================================
    # PART 2 — Excel cache validation (optional, requires xlwings+Excel)
    #
    # Reads the ORIGINAL source in Excel (live recalculation) and compares
    # against openpyxl's cached values from Part 1.  This validates that
    # the cached formula results openpyxl sees are up to date.
    #
    # The written file cannot be opened by Excel via AppleScript (openpyxl
    # strips data validation extensions), so formula recalculation of the
    # written file must be verified manually by opening it in Excel.
    # ==================================================================
    excel_checked = 0
    excel_mismatched = 0
    excel_available = False
    t_excel = 0.0

    print(f"\n  PART 2: Excel cache validation (live vs cached)")
    print(f"  {'─' * 40}")

    print(f"\n[5/5] Reading {source.name} via xlwings+Excel ...")
    t0 = time.time()
    orig_full = _xlwings_full_read(source, FIELD_MAP)
    t_excel = time.time() - t0

    if orig_full is None:
        print("      SKIPPED — xlwings or Excel not available")
        print("      Install xlwings and Excel for cache validation")
    else:
        excel_available = True
        orig_excel_flat = _flatten_for_compare(orig_full)
        n_excel = sum(len(v) for v in orig_excel_flat.values())
        print(f"      {n_excel} fields read via Excel, {t_excel:.1f}s")

        json_orig_excel = out_dir / f"{source_name}_orig_excel.json"
        json_orig_excel.write_text(
            json.dumps(orig_excel_flat, indent=2, default=str),
            encoding="utf-8"
        )

        # Compare Excel live values vs openpyxl cached values
        all_flat = _flatten_for_compare(data_all)
        excel_checked, excel_mismatched, excel_details = _compare_full(
            orig_excel_flat, all_flat)

        if excel_mismatched == 0:
            print(f"      *** All {excel_checked} fields match:"
                  f" openpyxl cache is fresh ***")
        else:
            print(f"      {excel_checked} fields compared,"
                  f" {excel_mismatched} stale cached values:")
            for d in excel_details[:20]:
                print(d)
            if len(excel_details) > 20:
                print(f"      ... and {len(excel_details) - 20} more")

        print(f"\n      Note: written file formula verification requires")
        print(f"      manually opening {written_path.name} in Excel.")

    # ==================================================================
    # Summary
    # ==================================================================
    total_t = part1_t + t_excel

    print(f"\n  {'═' * 40}")
    print(f"  Total time: {total_t:.1f}s")
    print(f"  Artifacts saved to {out_dir}/")

    return {
        "source": source.name,
        "template": template.name,
        "inputs_non_none": non_none_i,
        "all_non_none": non_none_a,
        "formulas_filtered": filtered,
        "cells_written": len(writes),
        "cells_verified": checked,
        "mismatches": mismatched,
        "excel_available": excel_available,
        "excel_checked": excel_checked,
        "excel_stale": excel_mismatched,
        "time_total": total_t,
    }


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    pairs = []
    for i in range(1, len(sys.argv), 2):
        source = Path(sys.argv[i])
        template = Path(sys.argv[i + 1]) if i + 1 < len(sys.argv) else None
        if template is None:
            print(f"Missing template for {source}")
            sys.exit(1)
        pairs.append((source, template))

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = ROOT / "records" / f"roundtrip_{stamp}"

    summaries = []
    for source, template in pairs:
        summaries.append(roundtrip(source, template, out_dir))

    if len(summaries) > 1:
        print(f"\n\n{'=' * 60}")
        print("  COMBINED SUMMARY")
        print(f"{'=' * 60}\n")
        for s in summaries:
            excel_info = ""
            if s["excel_available"]:
                excel_info = (f"  cache: {s['excel_checked']} checked,"
                              f" {s['excel_stale']} stale")
            print(f"  {s['source']:20s} → {s['template']:20s}"
                  f"  written: {s['cells_written']:>4}"
                  f"  verified: {s['cells_verified']:>4}"
                  f"  mismatches: {s['mismatches']:>3}"
                  f"{excel_info}"
                  f"  time: {s['time_total']:.1f}s")


if __name__ == "__main__":
    main()
