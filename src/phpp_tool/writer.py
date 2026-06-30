"""Write a building record into a PHPP workbook via openpyxl only.

No Excel installation required. The template is loaded twice:
  - data_only=True (read-only): resolve locator lookups (label searches,
    named ranges, entry-row detection) using cached cell values
  - data_only=False (writable): apply cell writes while preserving formulas

The writable workbook is saved to the output path. openpyxl's save strips
some Excel extensions (data validation rules, custom headers/footers), so
the output file may have degraded PHPP features.
"""

from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from phpp_tool.locators import (
    WsPair,
    classify_item,
    col_to_idx,
    field_col,
    find_row_in_col,
    is_entry_row_header,
    parse_cell_ref,
    prefer_si_sheet,
)
from phpp_tool.map_parser import parse_field_map

logger = logging.getLogger(__name__)


def write_phpp(
    record: dict[str, Any],
    template_path: str | Path,
    output_path: str | Path,
    field_map_path: str | Path = "phpp-field-mapping.md",
) -> list[tuple[str, str, int, Any]]:
    """Write a building record into a PHPP workbook template.

    Returns the list of writes performed as (sheet_name, col, row, value)
    tuples, for verification by callers.
    """
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    template_path = Path(template_path)
    output_path = Path(output_path)

    shutil.copy2(template_path, output_path)
    output_path.chmod(0o644)

    # Load template with data_only=True for label resolution (read-only)
    wb_labels = load_workbook(str(template_path), data_only=True)
    # Load output copy with data_only=False to preserve formulas (writable)
    wb_out = load_workbook(str(output_path), data_only=False)

    pending: list[tuple[str, str, int, Any]] = []

    try:
        field_map = parse_field_map(field_map_path)
        sheet_names = wb_labels.sheetnames
        total_writes = 0

        for ws_key, ws_data in record.items():
            if ws_data is None:
                continue
            ws_spec = field_map.get(ws_key)
            if ws_spec is None:
                logger.info("No field map entry for %r, skipping", ws_key)
                continue
            sheet_name = prefer_si_sheet(ws_spec["sheet_name"], sheet_names)
            if sheet_name not in sheet_names:
                logger.info("Sheet %r not in template, skipping %s",
                            sheet_name, ws_key)
                continue
            ws_labels = wb_labels[sheet_name]
            ws_out = wb_out[sheet_name]
            total_writes += _write_worksheet(
                ws_labels, ws_out, wb_labels, ws_spec, ws_data, pending)
    finally:
        wb_labels.close()

    # Apply writes and save
    for sheet_name, col, row, value in pending:
        wb_out[sheet_name].cell(row=row, column=col_to_idx(col), value=value)
    wb_out.save(str(output_path))
    wb_out.close()

    logger.info("Wrote %d cell values", total_writes)
    return pending


# ======================================================================
# Cell writing
# ======================================================================

def _write_cell(
    ws_out: Worksheet, col: str, row: int, value: Any,
    pending: list[tuple[str, str, int, Any]],
) -> bool:
    """Record a cell write for later persistence. Returns True if recorded."""
    if value is None:
        return False
    if isinstance(value, (dict, list)):
        return False
    pending.append((ws_out.title, col, row, value))
    return True


# ======================================================================
# Worksheet dispatch
# ======================================================================

def _write_worksheet(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    wb_labels: Workbook,
    ws_spec: dict[str, Any],
    ws_data: dict[str, Any],
    pending: list[tuple[str, str, int, Any]],
) -> int:
    """Write all mapped values into a single worksheet. Returns write count."""
    count = 0
    if ws_spec.get("fields"):
        count += _write_label_anchored(ws_labels, ws_out, ws_spec["fields"],
                                       ws_data, pending)

    for sec_name, sec_spec in ws_spec.get("sections", {}).items():
        sec_data = ws_data.get(sec_name)
        if sec_data is None:
            continue
        count += _write_section(ws_labels, ws_out, wb_labels, sec_spec,
                                sec_data, pending)
    return count


# --- Strategy 1: label-anchored fields ---

def _write_label_anchored(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    fields: dict[str, dict],
    data: dict[str, Any],
    pending: list[tuple[str, str, int, Any]],
) -> int:
    count = 0
    for field_name, spec in fields.items():
        value = data.get(field_name)
        if value is None or not spec.get("locator_string"):
            continue
        row = find_row_in_col(ws_labels, spec["locator_col"],
                              spec["locator_string"])
        if row is None:
            logger.warning("Label %r not found for write",
                           spec["locator_string"])
            continue
        if _write_cell(ws_out, spec["input_col"],
                       row + spec.get("row_offset", 0), value, pending):
            count += 1
    return count


# --- Section dispatch ---

def _write_section(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    wb_labels: Workbook,
    sec_spec: dict[str, Any],
    sec_data: Any,
    pending: list[tuple[str, str, int, Any]],
) -> int:
    has_header = "header_locator" in sec_spec
    has_entry = "entry_locator" in sec_spec
    has_col_fields = "column_fields" in sec_spec
    has_row_fields = "row_fields" in sec_spec
    has_items = "items" in sec_spec
    has_fields = "fields" in sec_spec

    items = sec_spec.get("items", {})
    entry_row_start = items.get(
        "entry_row_start") or items.get("entry_start_row")

    if has_header and has_entry and has_row_fields and not has_col_fields:
        return _write_row_offset(ws_labels, ws_out, sec_spec, sec_data,
                                 pending)
    if has_col_fields and has_row_fields:
        return _write_column_row(ws_labels, ws_out, sec_spec, sec_data,
                                 entry_row_start, pending)
    if has_header and has_entry and has_col_fields:
        return _write_block(ws_labels, ws_out, sec_spec, sec_data,
                            entry_row_start, pending)
    if has_col_fields and not has_header:
        return _write_static_column(ws_out, sec_spec, sec_data,
                                    entry_row_start, pending)
    if has_items:
        return _write_items(ws_labels, ws_out, wb_labels, items, sec_data,
                            pending)
    if has_fields:
        return _write_label_anchored(ws_labels, ws_out, sec_spec["fields"],
                                     sec_data, pending)
    return 0


# --- Strategy 2: repeating block ---

def _write_block(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    sec_spec: dict[str, Any],
    rows_data: list[dict[str, Any]],
    entry_row_start: int | None,
    pending: list[tuple[str, str, int, Any]],
) -> int:
    if not isinstance(rows_data, list):
        return 0

    if entry_row_start is None:
        entry_locator = sec_spec.get("entry_locator", {})
        entry_string = entry_locator.get("string", "")
        if entry_string:
            entry_col = entry_locator.get(
                "col", sec_spec.get("header_locator", {}).get("col", "A"))
            hdr = sec_spec.get("header_locator", {})
            hdr_row = 1
            if hdr.get("string"):
                hdr_found = find_row_in_col(
                    ws_labels, hdr.get("col", "A"), hdr["string"])
                if hdr_found is not None:
                    hdr_row = hdr_found
            found = find_row_in_col(
                ws_labels, entry_col, entry_string, start_from=hdr_row)
            if found is not None:
                col_fields = sec_spec.get("column_fields", {})
                if is_entry_row_header(ws_labels, found, col_fields):
                    entry_row_start = found + 1
                else:
                    entry_row_start = found

    has_row_metadata = all(
        "_row" in rd for rd in rows_data) if rows_data else False
    if entry_row_start is None and not has_row_metadata:
        logger.warning("Cannot determine block start row for write")
        return 0

    count = 0
    column_fields = sec_spec["column_fields"]
    for i, row_data in enumerate(rows_data):
        if "_row" in row_data:
            target_row = row_data["_row"]
        elif entry_row_start is not None:
            target_row = entry_row_start + i
        else:
            continue
        for field_name, field_spec in column_fields.items():
            if _write_cell(ws_out, field_col(field_spec), target_row,
                           row_data.get(field_name), pending):
                count += 1
    return count


# --- Row-offset section ---

def _write_row_offset(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    sec_spec: dict[str, Any],
    data: dict[str, Any],
    pending: list[tuple[str, str, int, Any]],
) -> int:
    if not isinstance(data, dict):
        return 0
    entry_loc = sec_spec["entry_locator"]
    anchor_row = find_row_in_col(ws_labels, entry_loc["col"],
                                 entry_loc["string"])
    if anchor_row is None:
        logger.warning("Entry locator %r not found for write",
                       entry_loc["string"])
        return 0
    input_col = sec_spec.get("items", {}).get("input_col_start", "J")
    count = 0
    for field_name, field_spec in sec_spec["row_fields"].items():
        value = data.get(field_name)
        if value is None:
            continue
        offset = field_spec.get("row_offset", field_spec.get("row", 0))
        if _write_cell(ws_out, input_col, anchor_row + offset, value,
                       pending):
            count += 1
    return count


# --- Column + row section ---

def _write_column_row(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    sec_spec: dict[str, Any],
    data: dict[str, Any],
    entry_row_start: int | None,
    pending: list[tuple[str, str, int, Any]],
) -> int:
    if not isinstance(data, dict):
        return 0
    if entry_row_start is None:
        entry_loc = sec_spec.get("entry_locator", {})
        if entry_loc:
            entry_row_start = find_row_in_col(
                ws_labels, entry_loc["col"], entry_loc["string"])
        if entry_row_start is None:
            return 0

    count = 0
    for col_name, col_spec in sec_spec["column_fields"].items():
        entity_data = data.get(col_name)
        if not isinstance(entity_data, dict):
            continue
        col_letter = field_col(col_spec)
        for field_name, field_spec in sec_spec["row_fields"].items():
            value = entity_data.get(field_name)
            if value is None:
                continue
            offset = field_spec.get("row_offset", field_spec.get("row", 0))
            if _write_cell(ws_out, col_letter, entry_row_start + offset,
                           value, pending):
                count += 1
    return count


# --- Static column section ---

def _write_static_column(
    ws_out: Worksheet,
    sec_spec: dict[str, Any],
    data: list[dict[str, Any]] | dict[str, Any],
    entry_row_start: int | None,
    pending: list[tuple[str, str, int, Any]],
) -> int:
    if not isinstance(data, list):
        return 0
    items = sec_spec.get("items", {})
    start_row = entry_row_start or items.get("entry_start_row")
    if start_row is None:
        return 0
    count = 0
    column_fields = sec_spec["column_fields"]
    for i, row_data in enumerate(data):
        target_row = row_data.get("_row", start_row + i)
        for field_name, field_spec in column_fields.items():
            if _write_cell(ws_out, field_col(field_spec), target_row,
                           row_data.get(field_name), pending):
                count += 1
    return count


# --- Items section ---

def _write_items(
    ws_labels: Worksheet,
    ws_out: Worksheet,
    wb_labels: Workbook,
    items_spec: dict[str, Any],
    data: dict[str, Any],
    pending: list[tuple[str, str, int, Any]],
) -> int:
    if not isinstance(data, dict):
        return 0
    count = 0
    for key, spec_value in items_spec.items():
        value = data.get(key)
        if value is None:
            continue
        if isinstance(spec_value, dict):
            if "col" in spec_value and "row" in spec_value:
                if _write_cell(ws_out, spec_value["col"],
                               spec_value["row"], value, pending):
                    count += 1
            continue
        kind = classify_item(key, spec_value)
        if kind == "address":
            col, row = parse_cell_ref(spec_value)
            if _write_cell(ws_out, col, row, value, pending):
                count += 1
        elif kind == "named_range":
            if _write_named_range(wb_labels, ws_out, spec_value, value,
                                  pending):
                count += 1
    return count


def _write_named_range(
    wb_labels: Workbook, ws_out: Worksheet, name: str, value: Any,
    pending: list[tuple[str, str, int, Any]],
) -> bool:
    """Resolve a named range via the labels workbook and record the write."""
    if value is None or isinstance(value, (dict, list)):
        return False
    try:
        defn = wb_labels.defined_names[name]
        for title, coord in defn.destinations:
            from openpyxl.utils import column_index_from_string
            from openpyxl.utils.cell import coordinate_from_string
            col_letter, row_num = coordinate_from_string(coord)
            pending.append((title, col_letter, row_num, value))
            return True
    except (KeyError, AttributeError):
        logger.warning("Named range %r not found for write", name)
    return False
