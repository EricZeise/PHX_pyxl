"""Resolve parsed locator specs against openpyxl worksheets.

Six addressing strategies:
  1. Label-anchored relative
  2. Header + entry block (repeating rows)
  3. Named ranges (German Excel defined names)
  4. Absolute address
  5. Column + row-offset within a block
  6. Fixed result rows/cols

All resolve functions accept a paired worksheet tuple (ws_vals, ws_fmls):
  - ws_vals: from load_workbook(data_only=True)  — cached values, label text
  - ws_fmls: from load_workbook(data_only=False)  — formula strings for detection

When skip_formulas is True, a cell whose ws_fmls value starts with "="
is returned as None.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

SPARSE_ROW_BREAK_THRESHOLD = 3

WsPair = tuple[Worksheet, Worksheet]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def norm(value: Any) -> str:
    """Normalize a label for comparison: NFKC, NBSP→space, strip, casefold."""
    if value is None:
        return ""
    s = str(value)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\xa0", " ")
    s = " ".join(s.split())
    return s.strip().casefold()


def resolve_sheet_name(sheet_name: str, sheet_names: list[str]) -> str | None:
    """Case-insensitively resolve *sheet_name* against real workbook sheets.

    Excel doesn't allow two sheets to coexist with names differing only by
    case, so matching case-insensitively is always safe -- it can't
    introduce ambiguity between two distinct real sheets. Returns the
    actual, correctly-cased name (needed for exact-case downstream lookups
    like ``wb[name]`` and the surgical writer's sheet map), or None if no
    sheet matches even case-insensitively.
    """
    target = sheet_name.casefold()
    for name in sheet_names:
        if name.casefold() == target:
            return name
    return None


def base_sheet_for_si_mirror(title: str, sheet_names: list[str]) -> str | None:
    """Return the base-tab name for a "<Name> SI"-suffixed sheet, or None.

    Returns None if *title* doesn't end in " SI" (case-insensitive) or no
    matching base sheet exists in *sheet_names*. Shared by the read-side
    SI-mirror-passthrough fallback and the write-side named-range redirect
    (both need to answer "what's the base tab for this SI sheet?").
    """
    if not title.lower().endswith(" si"):
        return None
    return resolve_sheet_name(title[: -len(" SI")], sheet_names)


def col_to_idx(col: str) -> int:
    """Convert column letters (A, ..., AA) to 1-based index."""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def field_col(spec: str | dict) -> str:
    """Extract column letter from a field spec (string or dict with 'column')."""
    return spec if isinstance(spec, str) else spec.get("column", "A")


def field_row_offset(spec: dict) -> int:
    """Extract a row offset from a field spec: 'row_offset', else 'row', else 0."""
    return spec.get("row_offset", spec.get("row", 0))


def resolve_entry_row_start(items: dict) -> int | None:
    """Look up a block's entry-start row from any of its recognized key aliases."""
    return (items.get("entry_row_start")
            or items.get("entry_start_row")
            or items.get("start_row"))


def _is_formula(ws_fmls: Worksheet, col_idx: int, row: int) -> bool:
    """Return True if the cell contains a formula (checked via data_only=False ws)."""
    val = ws_fmls.cell(row=row, column=col_idx).value
    return isinstance(val, str) and val.startswith("=")


def cell_value(
    ws_pair: WsPair, col: str, row: int, *, skip_formulas: bool = False,
) -> Any:
    """Read a single cell by column letter(s) and row number.

    When *skip_formulas* is True, returns None for formula cells so
    only designer-entered input values are captured.
    """
    ws_vals, ws_fmls = ws_pair
    col_idx = col_to_idx(col)
    if skip_formulas and _is_formula(ws_fmls, col_idx, row):
        return None
    return ws_vals.cell(row=row, column=col_idx).value


def find_row_in_col(
    ws_vals: Worksheet, col: str, needle: str, *,
    contains: bool = True, start_from: int = 1,
) -> int | None:
    """Return the first row where ``col``'s cell text matches *needle*.

    Uses the data_only=True worksheet for label text matching.
    """
    needle_n = norm(needle)
    if not needle_n:
        return None
    col_idx = col_to_idx(col)
    last_row = ws_vals.max_row or 1
    for row in range(start_from, last_row + 1):
        cell_val = ws_vals.cell(row=row, column=col_idx).value
        cell_n = norm(cell_val)
        if not cell_n:
            continue
        if contains:
            if needle_n in cell_n:
                return row
        else:
            if cell_n == needle_n:
                return row
    return None


def parse_cell_ref(ref: str) -> tuple[str, int]:
    """Split 'AB123' into ('AB', 123)."""
    m = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not m:
        raise ValueError(f"Invalid cell reference: {ref!r}")
    return m.group(1), int(m.group(2))


def is_header_row(row_data: dict[str, Any]) -> bool:
    """Return True if a block row looks like a header rather than data."""
    values = [v for k, v in row_data.items() if k != "_row" and v is not None]
    if not values:
        return False
    return all(isinstance(v, str) for v in values)


def is_entry_row_header(
    ws_vals: Worksheet, row: int, column_fields: dict[str, dict],
) -> bool:
    """Check if the entry locator row is a column header, not a data row."""
    row_data = {}
    for name, spec in column_fields.items():
        col_idx = col_to_idx(field_col(spec))
        row_data[name] = ws_vals.cell(row=row, column=col_idx).value
    return is_header_row(row_data)


# ---------------------------------------------------------------------------
# Strategy 1: Label-anchored relative
# ---------------------------------------------------------------------------

def is_label_anchored_formula(
    ws_pair: WsPair, locator_col: str, locator_string: str, input_col: str,
    row_offset: int = 0,
) -> bool | None:
    """Return whether a label-anchored field's target cell is a formula.

    Returns None if the label can't be found (caller should skip the
    input/output cross-check rather than treat this as a mismatch).
    """
    ws_vals, ws_fmls = ws_pair
    row = find_row_in_col(ws_vals, locator_col, locator_string)
    if row is None:
        return None
    col_idx = col_to_idx(input_col)
    return _is_formula(ws_fmls, col_idx, row + row_offset)


def resolve_label_anchored(
    ws_pair: WsPair,
    locator_col: str,
    locator_string: str,
    input_col: str,
    row_offset: int = 0,
    *,
    skip_formulas: bool = False,
) -> Any:
    """Find locator_string in locator_col, read input_col."""
    ws_vals, _ = ws_pair
    row = find_row_in_col(ws_vals, locator_col, locator_string)
    if row is None:
        logger.warning(
            "Label %r not found in column %s of sheet %r",
            locator_string, locator_col, ws_vals.title,
        )
        return None
    return cell_value(ws_pair, input_col, row + row_offset,
                      skip_formulas=skip_formulas)


# ---------------------------------------------------------------------------
# Strategy 2: Header + entry block (repeating rows)
# ---------------------------------------------------------------------------

_DEFAULT_END_MARKER = "Unhide additional rows"


def resolve_block(
    ws_pair: WsPair,
    header_locator: dict,
    entry_locator: dict,
    column_fields: dict[str, dict],
    *,
    end_marker: str = _DEFAULT_END_MARKER,
    entry_row_start: int | None = None,
    skip_formulas: bool = False,
) -> list[dict[str, Any]]:
    """Iterate a repeating block, returning one dict per data row.

    When *skip_formulas* is True, formula cells are returned as None.
    """
    ws_vals, ws_fmls = ws_pair
    entry_col = entry_locator.get("col") or header_locator.get("col") or "A"
    if not re.match(r"^[A-Za-z]{1,3}$", entry_col):
        # header_locator['col'] can hold a non-column placeholder (e.g. a
        # search string in the wrong field) when there's no real
        # entry_locator and the section is anchored purely by
        # entry_row_start -- fall back to a harmless default rather than
        # feeding a bogus column letter to col_to_idx().
        entry_col = "A"

    entry_string = entry_locator.get("string", "")

    def _discover_start_row() -> tuple[int | None, bool]:
        """Find the entry row by searching for the header + entry label.

        Returns (start_row, header_found) -- header_found distinguishes
        "header missing" from "entry label missing" for warning purposes.
        """
        if not entry_string:
            return None, False
        hdr_row = find_row_in_col(
            ws_vals, header_locator["col"], header_locator["string"]
        )
        if hdr_row is None:
            return None, False
        start_row_found = find_row_in_col(
            ws_vals, entry_col, entry_string, start_from=hdr_row,
        )
        if start_row_found is None:
            return None, True
        if is_entry_row_header(ws_vals, start_row_found, column_fields):
            return max(start_row_found + 1, hdr_row), True
        return max(start_row_found, hdr_row), True

    if entry_row_start is not None:
        start_row = entry_row_start
        # entry_row_start always wins (it's the authoritative override), but
        # cross-check it against the discoverable label position -- if the
        # two disagree, that's a sign the hardcoded row has drifted from the
        # workbook's actual layout, so surface it instead of staying silent.
        if entry_string:
            discovered, _ = _discover_start_row()
            if discovered is not None and discovered != entry_row_start:
                logger.warning(
                    "entry_row_start=%d for entry label %r in sheet %r "
                    "disagrees with the discovered row %d -- using "
                    "entry_row_start, but the field map may be stale",
                    entry_row_start, entry_string, ws_vals.title, discovered,
                )
    else:
        if not entry_string:
            return []
        discovered, header_found = _discover_start_row()
        if discovered is None:
            if not header_found:
                logger.warning(
                    "Block header %r not found in column %s",
                    header_locator["string"], header_locator["col"],
                )
            return []
        start_row = discovered

    last_row = ws_vals.max_row or 1
    if start_row > last_row:
        return []

    field_names = list(column_fields.keys())
    field_col_idxs = [col_to_idx(field_col(column_fields[f]))
                      for f in field_names]
    entry_col_idx = col_to_idx(entry_col)

    end_marker_n = norm(end_marker)
    results: list[dict[str, Any]] = []
    n_fields = len(column_fields)
    consecutive_sparse = 0

    for row_num in range(start_row, last_row + 1):
        entry_cell = ws_vals.cell(row=row_num, column=entry_col_idx)
        marker_val = norm(entry_cell.value)
        if end_marker_n and end_marker_n in marker_val:
            break

        # A bold entry-column cell is a section title (a totals row, or the
        # start of an unrelated table below that happens to reuse this
        # block's column layout), not more block data -- unlike
        # is_header_row() below, this doesn't depend on every mapped column
        # in the row happening to be a string, so it also catches rows that
        # mix a formula-cached number with text (e.g. a "Total ..." summary
        # row) and rows that otherwise look like valid data. Verified against
        # every header+entry block in both field-map versions: this is the
        # only bold row any block currently returns, so this can only ever
        # narrow (never break) existing results.
        if marker_val and entry_cell.font and entry_cell.font.bold:
            logger.debug(
                "Stopping block scan at row %d in sheet %r -- entry "
                "column is bold, signaling a new section", row_num,
                ws_vals.title,
            )
            break

        # Sparse/header detection always looks at raw (unfiltered) values,
        # so a row with real data isn't misclassified as blank just because
        # skip_formulas nulled out its formula-driven fields. row_data (the
        # returned dict) still applies the skip_formulas filter as before.
        row_data: dict[str, Any] = {"_row": row_num}
        raw_data: dict[str, Any] = {"_row": row_num}
        all_none = True
        for j, field_name in enumerate(field_names):
            col_idx = field_col_idxs[j]
            raw_val = ws_vals.cell(row=row_num, column=col_idx).value
            raw_data[field_name] = raw_val
            if raw_val is not None:
                all_none = False
            val = raw_val
            if skip_formulas and _is_formula(ws_fmls, col_idx, row_num):
                val = None
            row_data[field_name] = val

        if all_none:
            consecutive_sparse += 1
            if consecutive_sparse >= SPARSE_ROW_BREAK_THRESHOLD:
                break
            continue

        if is_header_row(raw_data):
            logger.debug("Skipping header row %d in sheet %r",
                         row_num, ws_vals.title)
            consecutive_sparse = 0
            continue

        non_row = [v for k, v in raw_data.items()
                   if k != "_row" and v is not None]
        has_string = any(isinstance(v, str) for v in non_row)
        if not has_string and len(non_row) <= max(n_fields // 3, 1):
            consecutive_sparse += 1
            if consecutive_sparse >= SPARSE_ROW_BREAK_THRESHOLD:
                break
            continue

        consecutive_sparse = 0
        results.append(row_data)

    return results


# ---------------------------------------------------------------------------
# Strategy 3: Named range
# ---------------------------------------------------------------------------

_NOT_FOUND = object()


def _resolve_si_mirror_passthrough(
    wb_vals, wb_fmls, title: str, coord: str,
) -> Any:
    """Fall back to the base-tab counterpart of a "<Name> SI" formula cell.

    Some Excel-internal defined names (e.g. `Klima_Region`, `Klima_Standort`)
    resolve to a cell on the "<Name> SI" mirror tab whose formula (typically
    `=IF(ISTEXT(<Name>!<coord>),<Name>!<coord>,"")`) just passes through the
    real designer input that lives at the same coordinate on the base tab --
    this is Excel's own defined-name table pointing there, not a choice the
    field map's `sheet_name` makes, so the existing per-version sheet_name
    fix (see locators.py module docstring / concern #24) can't route around
    it. Only applies when *title* actually ends in " SI" and the base tab's
    same-coordinate cell is itself not a formula; otherwise returns
    _NOT_FOUND so the caller keeps its existing behavior.
    """
    base_title = base_sheet_for_si_mirror(title, wb_vals.sheetnames)
    if base_title is None:
        return _NOT_FOUND
    base_fval = wb_fmls[base_title][coord].value
    if isinstance(base_fval, str) and base_fval.startswith("="):
        return _NOT_FOUND
    return wb_vals[base_title][coord].value


def resolve_named_range(
    wb_vals, wb_fmls, name: str, *, skip_formulas: bool = False,
) -> Any:
    """Resolve a German Excel defined name to its value.

    wb_vals and wb_fmls are openpyxl Workbook objects (data_only=True/False).
    """
    try:
        defn = wb_vals.defined_names[name]
    except KeyError:
        logger.warning("Named range %r not found in workbook", name)
        return None

    for title, coord in defn.destinations:
        ws_v = wb_vals[title]
        cell_v = ws_v[coord]
        if hasattr(cell_v, "__iter__") and not isinstance(cell_v, str):
            # Multi-cell destination -- PHPP defines several device-type-
            # name ranges broader than the single value they hold (e.g.
            # Kuehlgeraete_Kompressor_Umluft_Geraet -> K37:R37, every cell
            # blank except K37). Resolve to the top-left cell instead of
            # giving up entirely -- verified true for every multi-cell
            # named_range entry in the IP field map (3 of 17), each with
            # its real value (if any) only in the first cell.
            cell_v = cell_v[0][0]
        coord = cell_v.coordinate
        if skip_formulas:
            ws_f = wb_fmls[title]
            cell_f = ws_f[coord]
            fval = cell_f.value
            if isinstance(fval, str) and fval.startswith("="):
                base_value = _resolve_si_mirror_passthrough(
                    wb_vals, wb_fmls, title, coord)
                if base_value is not _NOT_FOUND:
                    return base_value
                return None
        return cell_v.value

    return None


# ---------------------------------------------------------------------------
# Strategy 4: Absolute address
# ---------------------------------------------------------------------------

def resolve_absolute(
    ws_pair: WsPair, address: str, *, skip_formulas: bool = False,
) -> Any:
    """Return the value at a fixed cell reference like 'C11'."""
    col, row = parse_cell_ref(address)
    return cell_value(ws_pair, col, row, skip_formulas=skip_formulas)


# ---------------------------------------------------------------------------
# Strategy 5: Column + row-offset within a block
# ---------------------------------------------------------------------------

def resolve_row_offset(
    ws_pair: WsPair,
    anchor_row: int,
    col: str,
    row_offset: int = 0,
    *,
    skip_formulas: bool = False,
) -> Any:
    """Return value at *col*, *anchor_row* + *row_offset*."""
    return cell_value(ws_pair, col, anchor_row + row_offset,
                      skip_formulas=skip_formulas)


# ---------------------------------------------------------------------------
# Strategy 6: Fixed result rows/cols
# ---------------------------------------------------------------------------

def resolve_fixed(
    ws_pair: WsPair, *, row: int, col: str, skip_formulas: bool = False,
) -> Any:
    """Read a fixed result location (typically formula outputs)."""
    return cell_value(ws_pair, col, row, skip_formulas=skip_formulas)
