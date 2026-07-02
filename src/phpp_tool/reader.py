"""Read a filled PHPP workbook into a nested dict using the field map.

Uses openpyxl only — no Excel installation required. The workbook is
loaded twice:
  - data_only=True:  cached values for reading and label searching
  - data_only=False: formula strings for input/formula classification

Only designer-entered input values are captured by default. Formula
cells are skipped so the JSON record contains only data that can be
meaningfully written back.
"""

from __future__ import annotations

import logging
import re
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from phpp_tool.locators import (
    WsPair,
    field_col,
    find_row_in_col,
    is_label_anchored_formula,
    resolve_absolute,
    resolve_block,
    resolve_fixed,
    resolve_label_anchored,
    resolve_named_range,
    resolve_row_offset,
    resolve_sheet_name,
)
from phpp_tool.map_parser import parse_field_map

logger = logging.getLogger(__name__)


def read_phpp(
    workbook_path: str | Path,
    field_map_path: str | Path = "phpp-field-mapping.md",
    *,
    skip_formulas: bool = True,
) -> dict[str, Any]:
    """Read a filled PHPP workbook into a nested dict.

    When *skip_formulas* is True (default), only designer-entered input
    values are captured.  Formula cells are returned as None so the JSON
    record contains only data that can be meaningfully written back.
    """
    warnings.filterwarnings("ignore", category=UserWarning)
    path = str(Path(workbook_path).resolve())

    wb_vals = load_workbook(path, data_only=True)
    wb_fmls = load_workbook(path, data_only=False)

    try:
        field_map = parse_field_map(field_map_path)
        sheet_names = wb_vals.sheetnames
        result: dict[str, Any] = {}

        for ws_key, ws_spec in field_map.items():
            sheet_name = resolve_sheet_name(ws_spec["sheet_name"], sheet_names)
            if sheet_name is None:
                logger.warning("Sheet %r not found, skipping %s",
                               ws_spec["sheet_name"], ws_key)
                continue
            ws_pair: WsPair = (wb_vals[sheet_name], wb_fmls[sheet_name])
            ws_result = _read_worksheet(
                ws_pair, wb_vals, wb_fmls, ws_spec,
                skip_formulas=skip_formulas)
            if ws_result:
                result[ws_key] = ws_result
    finally:
        wb_vals.close()
        wb_fmls.close()

    return result


def _read_worksheet(
    ws_pair: WsPair,
    wb_vals: Workbook,
    wb_fmls: Workbook,
    ws_spec: dict[str, Any],
    *, skip_formulas: bool = True,
) -> dict[str, Any]:
    """Extract all mapped values from a single worksheet."""
    ws_result: dict[str, Any] = {}

    if ws_spec.get("fields"):
        resolved = _read_label_anchored_fields(
            ws_pair, ws_spec["fields"], skip_formulas=skip_formulas)
        if resolved:
            ws_result.update(resolved)

    if ws_spec.get("config"):
        config_resolved = _read_config(
            ws_pair, wb_vals, wb_fmls, ws_spec["config"],
            ws_spec.get("config_kind", {}), skip_formulas=skip_formulas)
        if config_resolved:
            ws_result["_config"] = config_resolved

    for sec_name, sec_spec in ws_spec.get("sections", {}).items():
        sec_result = _read_section(
            ws_pair, wb_vals, wb_fmls, sec_spec, ws_spec.get("config", {}),
            skip_formulas=skip_formulas)
        if sec_result:
            ws_result[sec_name] = sec_result

    has_mapped_content = bool(
        ws_spec.get("fields") or ws_spec.get("config") or ws_spec.get("sections"))
    if has_mapped_content and not ws_result:
        logger.warning(
            "Worksheet %r has mapped fields/sections but resolved to no "
            "data at all -- check locators and skip_formulas interaction",
            ws_pair[0].title)

    return ws_result


# ---------------------------------------------------------------------------
# Top-level fields (Strategy 1: label-anchored)
# ---------------------------------------------------------------------------

_OPTION_CODE_RE = re.compile(r"^(\w+)-")


def _check_options(
    field_name: str, val: Any, options: dict[str, str] | None,
    sheet_name: str,
) -> None:
    """Warn if a resolved value's leading code isn't a documented option.

    The field map's ``options`` metadata (e.g. ``10``: 10-Passive house)
    is otherwise never cross-checked against what's actually in the
    workbook -- this makes drift between documented and real values
    visible instead of silently ignored.
    """
    if not options or not isinstance(val, str):
        return
    m = _OPTION_CODE_RE.match(val)
    code = m.group(1) if m else val
    if code not in options:
        logger.warning(
            "Field %r resolved to %r in sheet %r, which doesn't match any "
            "documented option code (%s) -- field map options may be stale",
            field_name, val, sheet_name, ", ".join(sorted(options)),
        )


def _check_io(
    field_name: str, spec: dict, ws_pair: WsPair, sheet_name: str,
) -> None:
    """Warn if a field's declared io tag disagrees with its actual formula status.

    The field map's io metadata (input/output) is otherwise never checked
    against the workbook -- this surfaces drift instead of silently
    trusting a tag that may no longer match the sheet's layout.
    """
    io = spec.get("io")
    if io is None:
        return
    is_formula = is_label_anchored_formula(
        ws_pair, spec["locator_col"], spec["locator_string"],
        spec["input_col"], spec.get("row_offset", 0),
    )
    if is_formula is None:
        return
    if io == "input" and is_formula:
        logger.warning(
            "Field %r in sheet %r is tagged (input) but its cell contains "
            "a formula -- field map may be stale",
            field_name, sheet_name,
        )
    elif io == "output" and not is_formula:
        logger.warning(
            "Field %r in sheet %r is tagged (output) but its cell is a "
            "literal value, not a formula -- field map may be stale",
            field_name, sheet_name,
        )


def _read_label_anchored_fields(
    ws_pair: WsPair, fields: dict[str, dict],
    *, skip_formulas: bool = True,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for field_name, spec in fields.items():
        if not spec.get("locator_string"):
            continue
        val = resolve_label_anchored(
            ws_pair,
            locator_col=spec["locator_col"],
            locator_string=spec["locator_string"],
            input_col=spec["input_col"],
            row_offset=spec.get("row_offset", 0),
            skip_formulas=skip_formulas,
        )
        _check_options(field_name, val, spec.get("options"), ws_pair[0].title)
        _check_io(field_name, spec, ws_pair, ws_pair[0].title)
        result[field_name] = val
    return result


# ---------------------------------------------------------------------------
# Config values
# ---------------------------------------------------------------------------

def _read_config(
    ws_pair: WsPair,
    wb_vals: Workbook,
    wb_fmls: Workbook,
    config: dict[str, Any],
    config_kind: dict[str, str],
    *, skip_formulas: bool = True,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in config.items():
        kind = config_kind.get(key, "literal")
        if kind == "address":
            result[key] = resolve_absolute(ws_pair, value,
                                           skip_formulas=skip_formulas)
        elif kind == "named_range":
            result[key] = resolve_named_range(wb_vals, wb_fmls, value,
                                              skip_formulas=skip_formulas)
        else:
            result[key] = value
    return result


# ---------------------------------------------------------------------------
# Section dispatch
# ---------------------------------------------------------------------------

def _read_section(
    ws_pair: WsPair,
    wb_vals: Workbook,
    wb_fmls: Workbook,
    sec_spec: dict[str, Any],
    ws_config: dict[str, Any],
    *,
    skip_formulas: bool = True,
) -> dict[str, Any] | list[dict[str, Any]] | None:
    """Read a single section, dispatching based on what keys are present."""
    has_header = "header_locator" in sec_spec
    has_entry = "entry_locator" in sec_spec
    has_col_fields = "column_fields" in sec_spec
    has_row_fields = "row_fields" in sec_spec
    has_items = "items" in sec_spec
    has_fields = "fields" in sec_spec
    has_appliance = "appliance_rows" in sec_spec

    items = sec_spec.get("items", {})
    entry_row_start = (items.get("entry_row_start")
                       or items.get("entry_start_row")
                       or items.get("start_row"))
    sf = skip_formulas

    if has_header and has_entry and has_row_fields and not has_col_fields:
        return _read_row_offset_section(ws_pair, sec_spec, skip_formulas=sf)
    if has_col_fields and has_row_fields:
        return _read_column_row_section(ws_pair, sec_spec, entry_row_start,
                                        skip_formulas=sf)
    if has_header and has_entry and has_col_fields:
        return _read_block_section(ws_pair, sec_spec, entry_row_start,
                                   skip_formulas=sf)
    if has_col_fields and not has_header:
        return _read_static_column_section(ws_pair, sec_spec, entry_row_start,
                                           skip_formulas=sf)
    if has_col_fields and entry_row_start is not None:
        return _read_block_section(ws_pair, sec_spec, entry_row_start,
                                   skip_formulas=sf)
    if has_appliance:
        return _read_appliance_section(sec_spec)
    if has_items:
        return _read_items_section(ws_pair, wb_vals, wb_fmls, items,
                                   sec_spec.get("items_kind", {}),
                                   skip_formulas=sf)
    if has_fields:
        return _read_label_anchored_fields(ws_pair, sec_spec["fields"],
                                           skip_formulas=sf)
    if has_header and not has_entry and not has_col_fields:
        return _read_header_only(ws_pair, sec_spec)
    return None


# ---------------------------------------------------------------------------
# Block patterns
# ---------------------------------------------------------------------------

def _read_block_section(
    ws_pair: WsPair,
    sec_spec: dict[str, Any],
    entry_row_start: int | None,
    *,
    skip_formulas: bool = True,
) -> list[dict[str, Any]]:
    return resolve_block(
        ws_pair,
        header_locator=sec_spec["header_locator"],
        entry_locator=sec_spec.get("entry_locator", {}),
        column_fields=sec_spec["column_fields"],
        entry_row_start=entry_row_start,
        skip_formulas=skip_formulas,
    )


def _read_row_offset_section(
    ws_pair: WsPair, sec_spec: dict[str, Any],
    *, skip_formulas: bool = True,
) -> dict[str, Any]:
    ws_vals, _ = ws_pair
    entry_loc = sec_spec["entry_locator"]
    anchor_row = find_row_in_col(ws_vals, entry_loc["col"], entry_loc["string"])
    if anchor_row is None:
        logger.warning("Entry locator %r not found", entry_loc["string"])
        return {}
    input_col = sec_spec.get("items", {}).get("input_col_start", "J")
    result: dict[str, Any] = {}
    for field_name, field_spec in sec_spec["row_fields"].items():
        offset = field_spec.get("row_offset", field_spec.get("row", 0))
        result[field_name] = resolve_row_offset(
            ws_pair, anchor_row, input_col, offset,
            skip_formulas=skip_formulas)
    return result


def _read_column_row_section(
    ws_pair: WsPair,
    sec_spec: dict[str, Any],
    entry_row_start: int | None,
    *, skip_formulas: bool = True,
) -> dict[str, Any]:
    ws_vals, _ = ws_pair
    entry_loc = sec_spec.get("entry_locator", {})
    if entry_row_start is None:
        if entry_loc:
            entry_row_start = find_row_in_col(
                ws_vals, entry_loc["col"], entry_loc["string"])
        if entry_row_start is None:
            return {}
    elif entry_loc:
        # entry_row_start always wins (it's the authoritative override), but
        # cross-check it against the discoverable label position -- if the
        # two disagree, that's a sign the hardcoded row has drifted from the
        # workbook's actual layout, so surface it instead of staying silent.
        discovered = find_row_in_col(ws_vals, entry_loc["col"], entry_loc["string"])
        if discovered is not None and discovered != entry_row_start:
            logger.warning(
                "entry_row_start=%d for entry label %r in sheet %r "
                "disagrees with the discovered row %d -- using "
                "entry_row_start, but the field map may be stale",
                entry_row_start, entry_loc["string"], ws_vals.title, discovered,
            )

    result: dict[str, Any] = {}
    for col_name, col_spec in sec_spec["column_fields"].items():
        col_letter = field_col(col_spec)
        entity: dict[str, Any] = {}
        for field_name, field_spec in sec_spec["row_fields"].items():
            offset = field_spec.get("row_offset", field_spec.get("row", 0))
            entity[field_name] = resolve_row_offset(
                ws_pair, entry_row_start, col_letter, offset,
                skip_formulas=skip_formulas)
        result[col_name] = entity
    return result


def _read_static_column_section(
    ws_pair: WsPair,
    sec_spec: dict[str, Any],
    entry_row_start: int | None,
    *, skip_formulas: bool = True,
) -> dict[str, Any] | list[dict[str, Any]]:
    ws_vals, _ = ws_pair
    items = sec_spec.get("items", {})
    start_row = entry_row_start or items.get("entry_start_row")

    if start_row is not None:
        rows: list[dict[str, Any]] = []
        last_row = ws_vals.max_row or 1
        for row in range(start_row, last_row + 1):
            row_data: dict[str, Any] = {"_row": row}
            all_none = True
            for field_name, field_spec in sec_spec["column_fields"].items():
                val = resolve_row_offset(ws_pair, row, field_col(field_spec), 0,
                                         skip_formulas=skip_formulas)
                row_data[field_name] = val
                if val is not None:
                    all_none = False
            if all_none:
                break
            rows.append(row_data)
        return rows

    return {"_columns": sec_spec["column_fields"]}


# ---------------------------------------------------------------------------
# Items-only sections
# ---------------------------------------------------------------------------

def _read_items_section(
    ws_pair: WsPair,
    wb_vals: Workbook,
    wb_fmls: Workbook,
    items: dict[str, Any],
    items_kind: dict[str, str],
    *, skip_formulas: bool = True,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in items.items():
        if isinstance(value, dict):
            if "col" in value and "row" in value:
                result[key] = resolve_fixed(
                    ws_pair, row=value["row"], col=value["col"],
                    skip_formulas=skip_formulas)
            else:
                result[key] = value
        else:
            kind = items_kind.get(key, "literal")
            if kind == "address":
                result[key] = resolve_absolute(ws_pair, value,
                                               skip_formulas=skip_formulas)
            elif kind == "named_range":
                result[key] = resolve_named_range(wb_vals, wb_fmls, value,
                                                  skip_formulas=skip_formulas)
            else:
                result[key] = value
    return result


# ---------------------------------------------------------------------------
# Appliance rows (Electricity) — stub
# ---------------------------------------------------------------------------

def _read_appliance_section(
    sec_spec: dict[str, Any],
) -> dict[str, Any]:
    """Return appliance row metadata from the Electricity sheet."""
    result: dict[str, Any] = {}
    for app_name, app_spec in sec_spec["appliance_rows"].items():
        data_row = app_spec.get("data_row")
        if data_row is None:
            continue
        entry: dict[str, Any] = {"data_row": data_row}
        if "selection_row" in app_spec:
            entry["selection_row"] = app_spec["selection_row"]
        if "options" in app_spec:
            entry["options"] = app_spec["options"]
        result[app_name] = entry
    return result


# ---------------------------------------------------------------------------
# Header-only sections
# ---------------------------------------------------------------------------

def _read_header_only(
    ws_pair: WsPair, sec_spec: dict[str, Any],
) -> dict[str, Any] | None:
    ws_vals, _ = ws_pair
    header = sec_spec["header_locator"]
    row = find_row_in_col(ws_vals, header["col"], header["string"])
    if row is None:
        return None
    return {"_header_row": row}
