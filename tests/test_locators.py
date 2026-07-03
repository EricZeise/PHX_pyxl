"""Tests for locators.py, in particular resolve_block's block-boundary logic."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from phpp_tool.locators import resolve_block


def _build_sheet(cells: dict[str, object], bold_rows: set[int]) -> Workbook:
    """Build a minimal workbook with a "Lighting"-style block.

    Row 10: header ("Lighting"). Row 11: entry locator ("Room / Zone").
    Rows 12+: data, as given by *cells* (col letter + row number -> value).
    Any row in *bold_rows* gets its entry-column (C) cell bolded.
    """
    wb = Workbook()
    ws = wb.active
    ws["C10"] = "Lighting"
    ws["C11"] = "Room / Zone"
    for ref, value in cells.items():
        ws[ref] = value
    for row in bold_rows:
        ws.cell(row=row, column=3).font = Font(bold=True)
    return wb


_COLUMN_FIELDS = {
    "room_zone_name": {"column": "C"},
    "net_floor_area": {"column": "D"},
}
_HEADER_LOCATOR = {"col": "C", "string": "Lighting"}
_ENTRY_LOCATOR = {"col": "C", "string": "Room / Zone"}


class TestResolveBlockBoldStop:
    def test_stops_at_bold_totals_row_instead_of_reading_past_it(self):
        wb = _build_sheet(
            cells={
                "C12": "Room A", "D12": 100,
                "C13": "Room B", "D13": 200,
                "C14": "Total floor area", "D14": 300,
                # An unrelated table below that reuses the same columns --
                # must never be reached once the bold row above is hit.
                "C15": "2- Laptop (standard)", "D15": 7,
            },
            bold_rows={14},
        )
        ws = wb.active
        rows = resolve_block(
            (ws, ws), _HEADER_LOCATOR, _ENTRY_LOCATOR, _COLUMN_FIELDS,
            entry_row_start=12,
        )

        assert [r["_row"] for r in rows] == [12, 13]
        assert rows[0]["room_zone_name"] == "Room A"
        assert rows[1]["room_zone_name"] == "Room B"

    def test_non_bold_rows_are_unaffected(self):
        wb = _build_sheet(
            cells={
                "C12": "Room A", "D12": 100,
                "C13": "Room B", "D13": 200,
                "C14": "Room C", "D14": 300,
            },
            bold_rows=set(),
        )
        ws = wb.active
        rows = resolve_block(
            (ws, ws), _HEADER_LOCATOR, _ENTRY_LOCATOR, _COLUMN_FIELDS,
            entry_row_start=12,
        )

        assert [r["_row"] for r in rows] == [12, 13, 14]

    def test_blank_bold_row_does_not_trigger_stop(self):
        # A bold cell with no value (leftover formatting) shouldn't count --
        # only a bold cell that actually carries content is a section title.
        wb = _build_sheet(
            cells={
                "C12": "Room A", "D12": 100,
                "C13": "Room B", "D13": 200,
            },
            bold_rows={14},
        )
        ws = wb.active
        rows = resolve_block(
            (ws, ws), _HEADER_LOCATOR, _ENTRY_LOCATOR, _COLUMN_FIELDS,
            entry_row_start=12,
        )

        assert [r["_row"] for r in rows] == [12, 13]
